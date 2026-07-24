import hashlib
import json
import tempfile
import unittest
import unicodedata
from pathlib import Path

import openpyxl

from certa.active_v1.dataset_adapter_v1 import (
    AITQAAdapterV1,
    AdapterValidationError,
    DuplicateTableIdentityError,
    HiTabAdapterV1,
    SSTQAZhAdapterV1,
    canonical_json_sha256,
    roundtrip_adapter_artifact,
)


def write_json(path, payload):
    path.write_text(
        json.dumps(payload, ensure_ascii=False, sort_keys=True),
        encoding="utf-8",
    )


def hitab_table():
    return {
        "title": "Example",
        "texts": [
            ["", "Metric"],
            ["Entity", "Value"],
            ["A", "4"],
            ["B", "2"],
        ],
        "top_root": {
            "row_index": -1,
            "column_index": -1,
            "children": [
                {
                    "row_index": 0,
                    "column_index": 1,
                    "children": [
                        {
                            "row_index": 1,
                            "column_index": 1,
                            "children": [],
                        }
                    ],
                }
            ],
        },
        "left_root": {
            "row_index": -1,
            "column_index": -1,
            "children": [
                {
                    "row_index": 2,
                    "column_index": 0,
                    "children": [],
                },
                {
                    "row_index": 3,
                    "column_index": 0,
                    "children": [],
                },
            ],
        },
        "merged_regions": [],
        "top_header_rows_num": 2,
        "left_header_columns_num": 1,
    }


def aitqa_raw_table():
    return {
        "id": "tab-1",
        "column_header": [
            ["Row group"],
            ["Year ended", "2020"],
            ["Year ended", "2021"],
        ],
        "row_header": [
            ["Revenue"],
            ["Cost"],
        ],
        "data": [
            ["4", "5"],
            ["2", "3"],
        ],
    }


def aitqa_flat_table():
    return [
        ["Row group", "Year ended", "Year ended"],
        ["", "2020", "2021"],
        ["Revenue", "4", "5"],
        ["Cost", "2", "3"],
    ]


class HiTabAdapterV1Tests(unittest.TestCase):
    def test_alias_is_provenance_not_a_path_and_tree_coordinates_are_normalized(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            write_json(root / "t-1.json", hitab_table())
            adapter = HiTabAdapterV1(root)

            self.assertEqual(adapter.discover()["table_file_count"], 1)
            self.assertEqual(tuple(adapter.index_tables()), ("t-1",))
            native = adapter.resolve_table(
                "t-1",
                runtime_record={"table_source": "statcan"},
            )
            artifact = adapter.canonicalize_table(native)
            adapter.validate_canonical_table(artifact)

            self.assertEqual(artifact["dataset"], "HiTab")
            self.assertEqual(
                set(artifact),
                {
                    "schema_version",
                    "dataset",
                    "table_id",
                    "adapter_id",
                    "source_identity",
                    "structure_summary",
                    "table_payload",
                },
            )
            self.assertEqual(
                artifact["source_identity"]["runtime_source_alias"],
                "statcan",
            )
            normalized = artifact["table_payload"]["graph_payload"]
            child = normalized["top_root"]["children"][0]
            self.assertEqual((child["row"], child["column"]), (0, 1))
            self.assertEqual((child["row_index"], child["column_index"]), (0, 1))

            smoke = roundtrip_adapter_artifact(
                artifact,
                question="What is the value for A?",
            )
            self.assertTrue(smoke["pass"], smoke)
            self.assertGreater(smoke["planner_schema_node_count"], 0)
            self.assertGreater(smoke["planner_schema_edge_count"], 0)
            self.assertEqual(smoke["grounding_state"], "UNIQUE")
            self.assertEqual(smoke["closure_outcome"], "UNIQUE_EXECUTABLE")

    def test_invalid_table_fails_without_poisoning_valid_table(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            write_json(root / "bad.json", {"texts": []})
            write_json(root / "good.json", hitab_table())
            adapter = HiTabAdapterV1(root)
            adapter.index_tables()

            with self.assertRaisesRegex(AdapterValidationError, "empty_table"):
                adapter.canonicalize_table(adapter.resolve_table("bad"))

            good = adapter.canonicalize_table(adapter.resolve_table("good"))
            adapter.validate_canonical_table(good)
            self.assertEqual(good["structure_summary"]["rows"], 4)

    def test_declared_span_extends_graph_grid_without_mutating_native_grid(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            table = hitab_table()
            table["merged_regions"] = [
                {
                    "first_row": 0,
                    "last_row": 0,
                    "first_column": 1,
                    "last_column": 3,
                }
            ]
            write_json(root / "wide-span.json", table)
            adapter = HiTabAdapterV1(root)

            artifact = adapter.canonicalize_table(
                adapter.resolve_table("wide-span"),
            )

            self.assertEqual(
                max(
                    len(row)
                    for row in artifact["table_payload"]["native_payload"][
                        "texts"
                    ]
                ),
                2,
            )
            self.assertEqual(artifact["structure_summary"]["columns"], 4)
            self.assertTrue(
                all(
                    len(row) == 4
                    for row in artifact["table_payload"]["graph_payload"][
                        "texts"
                    ]
                )
            )

    def test_formatted_numeric_surface_matches_executed_numeric_denotation(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            table = hitab_table()
            table["texts"][2][1] = "5,765"
            write_json(root / "formatted.json", table)
            adapter = HiTabAdapterV1(root)
            artifact = adapter.canonicalize_table(
                adapter.resolve_table("formatted"),
            )

            smoke = roundtrip_adapter_artifact(
                artifact,
                question="What is the formatted value for A?",
            )

            self.assertTrue(smoke["pass"], smoke)
            self.assertEqual(smoke["projected_answer"], "5,765")


class AITQAAdapterV1Tests(unittest.TestCase):
    def make_root(
        self,
        root,
        *,
        conflicting_duplicate=False,
        orphan_row_header=False,
    ):
        table = aitqa_raw_table()
        if orphan_row_header:
            table["row_header"].append(["Orphan native header"])
        raw = {
            "id": "q-1",
            "table_id": "tab-1",
            "question": "What is 2020 revenue?",
            "answers": ["4"],
            "table": table,
        }
        duplicate = dict(raw, id="q-2", answers=["5"])
        if conflicting_duplicate:
            duplicate = json.loads(json.dumps(duplicate))
            duplicate["table"]["data"][0][0] = "999"
        (root / "test_samples.jsonl").write_text(
            "\n".join(
                json.dumps(row, ensure_ascii=False)
                for row in (raw, duplicate)
            )
            + "\n",
            encoding="utf-8",
        )
        write_json(
            root / "aitqa_clean_questions.json",
            [
                {
                    "id": "q-1",
                    "table_id": "tab-1",
                    "question": "What is 2020 revenue?",
                    "answers": ["4"],
                    "table": aitqa_flat_table(),
                }
            ],
        )

    def test_raw_hierarchy_is_authoritative_and_clean_flattening_is_exact(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            self.make_root(root)
            adapter = AITQAAdapterV1(root)

            discovery = adapter.discover()
            self.assertEqual(discovery["raw_question_count"], 2)
            self.assertEqual(discovery["clean_question_count"], 1)
            self.assertEqual(tuple(adapter.index_tables()), ("tab-1",))
            artifact = adapter.canonicalize_table(
                adapter.resolve_table("tab-1"),
            )
            adapter.validate_canonical_table(artifact)

            self.assertEqual(
                artifact["table_payload"]["graph_payload"]["texts"],
                aitqa_flat_table(),
            )
            self.assertEqual(
                artifact["table_payload"]["native_payload"]["row_header"],
                [["Revenue"], ["Cost"]],
            )
            serialized = json.dumps(artifact, ensure_ascii=False)
            self.assertNotIn('"answers"', serialized)
            self.assertNotIn('"question"', serialized)

            smoke = roundtrip_adapter_artifact(
                artifact,
                question="What is 2020 revenue?",
            )
            self.assertTrue(smoke["pass"], smoke)
            self.assertEqual(smoke["grounding_state"], "UNIQUE")

    def test_conflicting_duplicate_table_identity_is_explicit(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            self.make_root(root, conflicting_duplicate=True)
            adapter = AITQAAdapterV1(root)
            with self.assertRaisesRegex(
                DuplicateTableIdentityError,
                "duplicate_table_payload_conflict:tab-1",
            ):
                adapter.index_tables()

    def test_orphan_native_row_header_is_preserved_but_not_fabricated_as_data(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            self.make_root(root, orphan_row_header=True)
            adapter = AITQAAdapterV1(root)

            adapter.discover()
            artifact = adapter.canonicalize_table(
                adapter.resolve_table("tab-1"),
            )

            self.assertEqual(
                len(
                    artifact["table_payload"]["native_payload"][
                        "row_header"
                    ]
                ),
                3,
            )
            self.assertEqual(
                artifact["table_payload"]["graph_payload"]["texts"],
                aitqa_flat_table(),
            )
            self.assertEqual(
                artifact["table_payload"]["graph_payload"]["aitqa_meta"][
                    "orphan_row_header_count"
                ],
                1,
            )


class SSTQAZhAdapterV1Tests(unittest.TestCase):
    def make_workbook(self, root):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "预算表"
        sheet.merge_cells("A1:C1")
        sheet["A1"] = "年度预算"
        sheet.append(["部门", "2020", "2021"])
        sheet.append(["研发", 4, 5])
        sheet.append(["销售", 2, 3])
        empty = workbook.create_sheet("空表")
        path = root / "49.xlsx"
        workbook.save(path)
        return path, empty.title

    def test_sheet_identity_active_alias_unicode_and_exact_grounding(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            path, _ = self.make_workbook(root)
            adapter = SSTQAZhAdapterV1(root)

            discovery = adapter.discover()
            self.assertEqual(discovery["workbook_count"], 1)
            self.assertEqual(discovery["sheet_count"], 2)
            index = adapter.index_tables()
            title_digest = hashlib.sha256(
                unicodedata.normalize("NFC", "预算表").encode("utf-8")
            ).hexdigest()[:16]
            sheet_id = f"sstqa_zh:49:sheet:000:{title_digest}"
            self.assertIn(sheet_id, index)
            self.assertEqual(adapter.resolve_table("49").table_id, sheet_id)
            self.assertEqual(
                adapter.resolve_table(sheet_id).source_path,
                path,
            )

            artifact = adapter.canonicalize_table(
                adapter.resolve_table(sheet_id),
            )
            adapter.validate_canonical_table(artifact)
            self.assertEqual(
                artifact["source_identity"]["sheet_title"],
                "预算表",
            )
            self.assertIn(
                "年度预算",
                json.dumps(artifact, ensure_ascii=False),
            )
            projection = artifact["table_payload"]["graph_payload"][
                "sstqa_zh_meta"
            ]["structural_projection"]
            self.assertEqual(
                projection["algorithm"],
                "minimal_exact_groundable_header_bands_v1",
            )

            smoke = roundtrip_adapter_artifact(
                artifact,
                question="研发部门的2020预算是多少？",
            )
            self.assertTrue(smoke["pass"], smoke)
            self.assertEqual(smoke["grounding_state"], "UNIQUE")
            self.assertEqual(smoke["closure_outcome"], "UNIQUE_EXECUTABLE")
            self.assertEqual(
                smoke["native_coordinate"],
                artifact["table_payload"]["native_payload"]["sheet"][
                    "cells"
                ][smoke["graph_row"]][smoke["graph_column"]]["coordinate"],
            )

            second = adapter.canonicalize_table(
                adapter.resolve_table(sheet_id),
            )
            self.assertEqual(
                canonical_json_sha256(artifact),
                canonical_json_sha256(second),
            )

    def test_empty_sheet_fails_locally_and_active_sheet_still_passes(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            self.make_workbook(root)
            adapter = SSTQAZhAdapterV1(root)
            index = adapter.index_tables()
            empty_id = next(
                table_id
                for table_id, row in index.items()
                if row.source_identity["sheet_title"] == "空表"
            )

            with self.assertRaisesRegex(AdapterValidationError, "empty_sheet"):
                adapter.canonicalize_table(adapter.resolve_table(empty_id))

            active = adapter.canonicalize_table(adapter.resolve_table("49"))
            adapter.validate_canonical_table(active)
            self.assertEqual(active["dataset"], "sstqa_zh")


if __name__ == "__main__":
    unittest.main()
