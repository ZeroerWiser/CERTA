import hashlib
import json
import tempfile
import unittest
from pathlib import Path

import openpyxl

from tools.certa_final_adapter_materialize import (
    OutputRootExistsError,
    materialize_adapter_stage,
)

from tests.active_v1.test_dataset_adapter_v1 import (
    aitqa_flat_table,
    aitqa_raw_table,
    hitab_table,
    write_json,
)


REPO = Path("/home/hsh/ME/Table/EMNLP2026/CERTA")
PACK = Path(
    "/home/hsh/ME/Table/EMNLP2026/certa_goal_packs/"
    "CERTA_FINAL_MULTI_DATASET_ADAPTER_AND_METHOD_COMPLETION_PACK"
)


def write_jsonl(path, rows):
    path.write_text(
        "".join(
            json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n"
            for row in rows
        ),
        encoding="utf-8",
    )


def sha256(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


class AdapterMaterializerV1Tests(unittest.TestCase):
    def make_sources(self, root):
        dataset = root / "dataset"
        hitab = dataset / "hitab"
        tables = hitab / "tables" / "raw"
        tables.mkdir(parents=True)
        for table_id in ("val-table", "hold-table"):
            write_json(tables / f"{table_id}.json", hitab_table())
        write_jsonl(
            hitab / "test_samples_clean.jsonl",
            [
                {
                    "id": "public-q",
                    "table_id": "val-table",
                    "question": "What is the value for A?",
                    "table_source": "statcan",
                    "answer": ["4"],
                    "answer_formulas": [],
                }
            ],
        )

        aitqa = dataset / "AIT-QA"
        aitqa.mkdir()
        raw = {
            "id": "ait-q",
            "table_id": "tab-1",
            "question": "What is 2020 revenue?",
            "answers": ["4"],
            "table": aitqa_raw_table(),
        }
        write_jsonl(aitqa / "test_samples.jsonl", [raw])
        write_json(
            aitqa / "aitqa_clean_questions.json",
            [
                {
                    "id": "ait-q",
                    "table_id": "tab-1",
                    "question": "What is 2020 revenue?",
                    "answers": ["4"],
                    "table": aitqa_flat_table(),
                }
            ],
        )

        sstqa = dataset / "SSTQA-zh"
        workbook_root = sstqa / "table"
        workbook_root.mkdir(parents=True)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "预算表"
        sheet.merge_cells("A1:C1")
        sheet["A1"] = "年度预算"
        sheet.append(["部门", "2020", "2021"])
        sheet.append(["研发", 4, 5])
        workbook.create_sheet("空表")
        workbook.save(workbook_root / "49.xlsx")
        write_jsonl(
            sstqa / "test.jsonl",
            [
                {
                    "id": "sst-q",
                    "table_id": 49,
                    "query": "研发部门的2020预算是多少？",
                    "label": "4",
                }
            ],
        )

        development = root / "development"
        development.mkdir()
        write_jsonl(
            development / "development_runtime.jsonl",
            [
                {
                    "dataset": "hitab",
                    "id": "dev-q",
                    "question": "What is the value for A?",
                    "table_id": "val-table",
                    "table_source": "statcan",
                }
            ],
        )
        write_json(
            development / "DEVELOPMENT_MANIFEST.json",
            {"schema_version": "test_development_manifest"},
        )

        strict = root / "strict"
        strict.mkdir()
        write_jsonl(
            strict / "fresh_validation_runtime_v2.jsonl",
            [
                {
                    "dataset": "HiTab",
                    "id": "val-q",
                    "question": "What is the value for A?",
                    "table_id": "val-table",
                    "table_source": "statcan",
                }
            ],
        )
        write_jsonl(
            strict / "fresh_holdout_runtime_v2.jsonl",
            [
                {
                    "dataset": "HiTab",
                    "id": "hold-q",
                    "question": "What is the value for B?",
                    "table_id": "hold-table",
                    "table_source": "totto",
                }
            ],
        )
        write_json(
            strict / "STRICT_FRESH_COHORT_MANIFEST.json",
            {"schema_version": "test_strict_manifest"},
        )
        return dataset, development, strict

    def test_atomic_materialization_has_exact_runtimes_and_verified_artifacts(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            dataset, development, strict = self.make_sources(root)
            output = root / "new-output"

            result = materialize_adapter_stage(
                repo=REPO,
                pack=PACK,
                dataset_root=dataset,
                development=development,
                strict_v2=strict,
                output=output,
                sentinel_ids={
                    "hitab": ["val-table"],
                    "aitqa": ["tab-1"],
                    "sstqa_zh": ["49"],
                },
                minimum_sentinels=1,
            )

            self.assertEqual(result["status"], "PASS")
            self.assertTrue(output.is_dir())
            for relative in (
                "intake/repository_runtime_identity.json",
                "data/DATASET_ROOT_BINDINGS.json",
                "data/DATASET_DISCOVERY_AUDIT.md",
                "data/hitab/TABLE_INDEX.jsonl",
                "data/aitqa/TABLE_INDEX.jsonl",
                "data/sstqa_zh/TABLE_INDEX.jsonl",
                "data/hitab/ADAPTER_SENTINEL_REPORT.json",
                "data/aitqa/ADAPTER_SENTINEL_REPORT.json",
                "data/sstqa_zh/ADAPTER_SENTINEL_REPORT.json",
                "data/hitab/CANONICAL_TABLE_MANIFEST.json",
                "data/hitab/validation_runtime_v3.jsonl",
                "data/hitab/holdout_runtime_v3.jsonl",
            ):
                self.assertTrue((output / relative).is_file(), relative)

            validation = json.loads(
                (output / "data/hitab/validation_runtime_v3.jsonl")
                .read_text(encoding="utf-8")
                .strip()
            )
            self.assertEqual(
                set(validation),
                {
                    "dataset",
                    "id",
                    "question",
                    "table_id",
                    "table_artifact",
                    "table_artifact_sha256",
                },
            )
            self.assertNotIn("table_source", validation)
            artifact = (
                output
                / "data/hitab/canonical_tables"
                / validation["table_artifact"]
            )
            self.assertEqual(sha256(artifact), validation["table_artifact_sha256"])
            self.assertEqual(json.loads(artifact.read_text())["dataset"], "HiTab")
            hitab_report = json.loads(
                (output / "data/hitab/ADAPTER_SENTINEL_REPORT.json").read_text()
            )
            sentinel_artifact = json.loads(
                (
                    output
                    / hitab_report["sentinels"][0]["canonical_artifact"]
                ).read_text()
            )
            self.assertEqual(
                sentinel_artifact["source_identity"]["runtime_source_alias"],
                "statcan",
            )
            self.assertEqual(
                hitab_report["grounding_smoke_scope"],
                "STRUCTURAL_LOOKUP_FIXTURE_NOT_QUESTION_ANSWER",
            )

            bindings = json.loads(
                (output / "data/DATASET_ROOT_BINDINGS.json").read_text()
            )
            self.assertEqual(
                bindings["operator_authoritative_dataset_root"],
                str(dataset.resolve()),
            )
            self.assertTrue(
                bindings["cohort_resolution"]["validation"]["all_resolved"]
            )
            self.assertTrue(
                bindings["cohort_resolution"]["holdout"]["all_resolved"]
            )
            self.assertEqual(
                bindings["cohort_resolution"]["validation_holdout_id_overlap"],
                [],
            )
            self.assertEqual(
                bindings["cohort_resolution"]["validation_holdout_table_overlap"],
                [],
            )
            self.assertEqual(
                len(bindings["native_question_container_access"]),
                3,
            )
            self.assertTrue(all(
                item["container_read"]
                and item["label_values_used_for_inference"] is False
                for item in bindings["native_question_container_access"]
            ))
            identity = json.loads(
                (output / "intake/repository_runtime_identity.json").read_text()
            )
            self.assertEqual(
                len(identity["native_label_bearing_containers_read"]),
                3,
            )
            for dataset_name in ("hitab", "aitqa", "sstqa_zh"):
                report = json.loads(
                    (
                        output
                        / f"data/{dataset_name}/ADAPTER_SENTINEL_REPORT.json"
                    ).read_text()
                )
                self.assertTrue(report["pass"], report)
                self.assertTrue(report["negative_isolation"]["pass"])

            with self.assertRaises(OutputRootExistsError):
                materialize_adapter_stage(
                    repo=REPO,
                    pack=PACK,
                    dataset_root=dataset,
                    development=development,
                    strict_v2=strict,
                    output=output,
                    sentinel_ids={
                        "hitab": ["val-table"],
                        "aitqa": ["tab-1"],
                        "sstqa_zh": ["49"],
                    },
                    minimum_sentinels=1,
                )


if __name__ == "__main__":
    unittest.main()
