"""Versioned, fail-closed dataset adapters for CERTA Active V1."""

from __future__ import annotations

import hashlib
import json
import unicodedata
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Optional, Sequence, Tuple

import openpyxl

from graph_builder import EdgeType, HCEG, NodeType, build_hceg

from certa.datasets.sstqa_zh import convert_sstqa_workbook
from certa.grounding.plan_closure import ClosureOutcome, build_plan_closure
from certa.grounding.structural_resolvers import (
    ResolutionState,
    resolve_atomic_operand,
)
from certa.planner.schema_view import (
    build_canonical_structural_group_catalog,
    build_proposal_blind_planner_view,
)
from certa.reproducibility.canonical_json import canonical_json, canonical_json_hash


CANONICAL_TABLE_SCHEMA_VERSION = "certa_canonical_table_v1"
SST_STRUCTURAL_PROJECTION_VERSION = "minimal_exact_groundable_header_bands_v1"
_CANONICAL_KEYS = {
    "schema_version",
    "dataset",
    "table_id",
    "adapter_id",
    "source_identity",
    "structure_summary",
    "table_payload",
}
_FORBIDDEN_LABEL_KEYS = {
    "answer",
    "answers",
    "correct",
    "correctness",
    "gold",
    "gold_answer",
    "label",
    "labels",
    "question",
}


class DatasetAdapterError(ValueError):
    """Base class for deterministic adapter failures."""


class AdapterDiscoveryError(DatasetAdapterError):
    """Raised when an authoritative native source cannot be discovered."""


class AdapterValidationError(DatasetAdapterError):
    """Raised when a native or canonical table violates the adapter contract."""


class DuplicateTableIdentityError(AdapterValidationError):
    """Raised when one table identity maps to conflicting native payloads."""


class TableResolutionError(DatasetAdapterError):
    """Raised when a requested table identity does not resolve exactly."""


@dataclass(frozen=True)
class TableIndexEntry:
    dataset: str
    table_id: str
    source_path: Path
    source_sha256: str
    source_identity: Dict[str, Any]
    payload_sha256: str = ""

    def to_record(self) -> Dict[str, Any]:
        return {
            "dataset": self.dataset,
            "table_id": self.table_id,
            "source_path": str(self.source_path),
            "source_sha256": self.source_sha256,
            "source_identity": dict(self.source_identity),
            "payload_sha256": self.payload_sha256,
        }


@dataclass(frozen=True)
class ResolvedNativeTable:
    dataset: str
    table_id: str
    source_path: Path
    source_sha256: str
    source_identity: Dict[str, Any]
    native_payload: Dict[str, Any]


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def canonical_json_sha256(value: Any) -> str:
    return canonical_json_hash(value)


def _root_node() -> Dict[str, Any]:
    return {
        "row": -1,
        "column": -1,
        "row_index": -1,
        "column_index": -1,
        "children": [],
    }


def _tree_node(
    row: int,
    column: int,
    children: Optional[Sequence[Mapping[str, Any]]] = None,
) -> Dict[str, Any]:
    return {
        "row": int(row),
        "column": int(column),
        "row_index": int(row),
        "column_index": int(column),
        "children": [dict(child) for child in (children or ())],
    }


def _axis_roots(
    rows: int,
    columns: int,
    top_header_rows: int,
    left_header_columns: int,
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    top_root = _root_node()
    left_root = _root_node()
    for column in range(left_header_columns, columns):
        parent: Optional[Dict[str, Any]] = None
        for row in range(top_header_rows - 1, -1, -1):
            parent = _tree_node(row, column, [parent] if parent else ())
        if parent is not None:
            top_root["children"].append(parent)
    for row in range(top_header_rows, rows):
        parent = None
        for column in range(left_header_columns - 1, -1, -1):
            parent = _tree_node(row, column, [parent] if parent else ())
        if parent is not None:
            left_root["children"].append(parent)
    return top_root, left_root


def _normalize_tree_coordinates(value: Any) -> Any:
    if not isinstance(value, Mapping):
        return value
    normalized = {
        str(key): _normalize_tree_coordinates(child)
        for key, child in value.items()
        if key != "children"
    }
    row = normalized.get("row", normalized.get("row_index", -1))
    column = normalized.get("column", normalized.get("column_index", -1))
    normalized["row"] = int(row)
    normalized["column"] = int(column)
    normalized.setdefault("row_index", int(row))
    normalized.setdefault("column_index", int(column))
    children = value.get("children") or []
    if not isinstance(children, list):
        raise AdapterValidationError("header_tree_children_not_list")
    normalized["children"] = [
        _normalize_tree_coordinates(child)
        for child in children
    ]
    return normalized


def _maximum_tree_coordinate(value: Any) -> Tuple[int, int]:
    if not isinstance(value, Mapping):
        return -1, -1
    row = int(value.get("row", value.get("row_index", -1)))
    column = int(
        value.get("column", value.get("column_index", -1))
    )
    for child in value.get("children") or []:
        child_row, child_column = _maximum_tree_coordinate(child)
        row = max(row, child_row)
        column = max(column, child_column)
    return row, column


def _pad_hitab_graph_grid(
    payload: Mapping[str, Any],
    texts: Sequence[Sequence[str]],
) -> list[list[str]]:
    if not texts or max((len(row) for row in texts), default=0) <= 0:
        raise AdapterValidationError("empty_table")
    required_rows = len(texts)
    required_columns = max(len(row) for row in texts)
    for region in payload.get("merged_regions") or []:
        if not isinstance(region, Mapping):
            raise AdapterValidationError("merged_region_not_object")
        try:
            required_rows = max(
                required_rows,
                int(region["last_row"]) + 1,
            )
            required_columns = max(
                required_columns,
                int(region["last_column"]) + 1,
            )
        except (KeyError, TypeError, ValueError) as exc:
            raise AdapterValidationError("merged_region_invalid") from exc
    for tree_name in ("top_root", "left_root"):
        row, column = _maximum_tree_coordinate(payload.get(tree_name))
        required_rows = max(required_rows, row + 1)
        required_columns = max(required_columns, column + 1)
    padded = [
        list(row) + [""] * (required_columns - len(row))
        for row in texts
    ]
    padded.extend(
        [[""] * required_columns for _ in range(required_rows - len(padded))]
    )
    return padded


def _as_text_rows(value: Any) -> list[list[str]]:
    if not isinstance(value, list):
        raise AdapterValidationError("table_rows_not_list")
    rows = []
    for row in value:
        if not isinstance(row, list):
            raise AdapterValidationError("table_row_not_list")
        rows.append(["" if item is None else str(item) for item in row])
    return rows


def _rectangular_rows(rows: Sequence[Sequence[Any]]) -> Tuple[int, int]:
    if not rows:
        raise AdapterValidationError("empty_table")
    columns = max((len(row) for row in rows), default=0)
    if columns <= 0:
        raise AdapterValidationError("empty_table")
    if any(len(row) != columns for row in rows):
        raise AdapterValidationError("ragged_table")
    return len(rows), columns


def _build_artifact(
    *,
    dataset: str,
    table_id: str,
    adapter_id: str,
    source_identity: Mapping[str, Any],
    native_payload: Mapping[str, Any],
    graph_payload: Mapping[str, Any],
) -> Dict[str, Any]:
    texts = _as_text_rows(graph_payload.get("texts"))
    rows, columns = _rectangular_rows(texts)
    merged_regions = list(graph_payload.get("merged_regions") or [])
    return {
        "schema_version": CANONICAL_TABLE_SCHEMA_VERSION,
        "dataset": dataset,
        "table_id": str(table_id),
        "adapter_id": adapter_id,
        "source_identity": dict(source_identity),
        "structure_summary": {
            "rows": rows,
            "columns": columns,
            "merged_region_count": len(merged_regions),
            "top_header_rows": int(
                graph_payload.get("top_header_rows_num", 0)
            ),
            "left_header_columns": int(
                graph_payload.get("left_header_columns_num", 0)
            ),
        },
        "table_payload": {
            "native_payload": dict(native_payload),
            "graph_payload": dict(graph_payload),
        },
    }


def _walk_keys(value: Any) -> Iterable[str]:
    if isinstance(value, Mapping):
        for key, child in value.items():
            yield str(key)
            yield from _walk_keys(child)
    elif isinstance(value, list):
        for child in value:
            yield from _walk_keys(child)


def _validate_tree(
    value: Any,
    *,
    rows: int,
    columns: int,
    name: str,
) -> None:
    if not isinstance(value, Mapping):
        raise AdapterValidationError(f"{name}_not_object")
    row = value.get("row")
    column = value.get("column")
    if not isinstance(row, int) or not isinstance(column, int):
        raise AdapterValidationError(f"{name}_coordinate_not_integer")
    if row < -1 or row >= rows or column < -1 or column >= columns:
        raise AdapterValidationError(f"{name}_coordinate_out_of_bounds")
    children = value.get("children")
    if not isinstance(children, list):
        raise AdapterValidationError(f"{name}_children_not_list")
    for child in children:
        _validate_tree(child, rows=rows, columns=columns, name=name)


def _validate_artifact(
    artifact: Mapping[str, Any],
    *,
    expected_dataset: Optional[str] = None,
    expected_adapter_id: Optional[str] = None,
) -> None:
    if not isinstance(artifact, Mapping):
        raise AdapterValidationError("canonical_artifact_not_object")
    if set(artifact) != _CANONICAL_KEYS:
        raise AdapterValidationError("canonical_artifact_key_mismatch")
    if artifact.get("schema_version") != CANONICAL_TABLE_SCHEMA_VERSION:
        raise AdapterValidationError("canonical_schema_version_mismatch")
    dataset = str(artifact.get("dataset") or "")
    table_id = str(artifact.get("table_id") or "")
    adapter_id = str(artifact.get("adapter_id") or "")
    if not dataset or not table_id or not adapter_id:
        raise AdapterValidationError("canonical_identity_empty")
    if expected_dataset is not None and dataset != expected_dataset:
        raise AdapterValidationError("canonical_dataset_mismatch")
    if expected_adapter_id is not None and adapter_id != expected_adapter_id:
        raise AdapterValidationError("canonical_adapter_id_mismatch")
    if not isinstance(artifact.get("source_identity"), Mapping):
        raise AdapterValidationError("canonical_source_identity_not_object")
    table_payload = artifact.get("table_payload")
    if not isinstance(table_payload, Mapping):
        raise AdapterValidationError("canonical_table_payload_not_object")
    native_payload = table_payload.get("native_payload")
    graph_payload = table_payload.get("graph_payload")
    if not isinstance(native_payload, Mapping):
        raise AdapterValidationError("native_payload_not_object")
    if not isinstance(graph_payload, Mapping):
        raise AdapterValidationError("graph_payload_not_object")
    forbidden = sorted(
        {
            key
            for key in _walk_keys(native_payload)
            if key.strip().lower() in _FORBIDDEN_LABEL_KEYS
        }
    )
    if forbidden:
        raise AdapterValidationError(
            f"label_field_in_native_payload:{','.join(forbidden)}"
        )
    texts = _as_text_rows(graph_payload.get("texts"))
    rows, columns = _rectangular_rows(texts)
    summary = artifact.get("structure_summary")
    if not isinstance(summary, Mapping):
        raise AdapterValidationError("structure_summary_not_object")
    if summary.get("rows") != rows or summary.get("columns") != columns:
        raise AdapterValidationError("structure_summary_shape_mismatch")
    top = graph_payload.get("top_header_rows_num")
    left = graph_payload.get("left_header_columns_num")
    if not isinstance(top, int) or not isinstance(left, int):
        raise AdapterValidationError("header_band_not_integer")
    if top <= 0 or top >= rows or left <= 0 or left >= columns:
        raise AdapterValidationError("header_band_has_no_data_region")
    merged_regions = graph_payload.get("merged_regions")
    if not isinstance(merged_regions, list):
        raise AdapterValidationError("merged_regions_not_list")
    if summary.get("merged_region_count") != len(merged_regions):
        raise AdapterValidationError("merged_region_count_mismatch")
    for region in merged_regions:
        if not isinstance(region, Mapping):
            raise AdapterValidationError("merged_region_not_object")
        try:
            first_row = int(region["first_row"])
            last_row = int(region["last_row"])
            first_column = int(region["first_column"])
            last_column = int(region["last_column"])
        except (KeyError, TypeError, ValueError) as exc:
            raise AdapterValidationError("merged_region_invalid") from exc
        if not (
            0 <= first_row <= last_row < rows
            and 0 <= first_column <= last_column < columns
        ):
            raise AdapterValidationError("merged_region_out_of_bounds")
    _validate_tree(
        graph_payload.get("top_root"),
        rows=rows,
        columns=columns,
        name="top_root",
    )
    _validate_tree(
        graph_payload.get("left_root"),
        rows=rows,
        columns=columns,
        name="left_root",
    )
    try:
        serialized = canonical_json(artifact)
        if canonical_json(json.loads(serialized)) != serialized:
            raise AdapterValidationError("canonical_roundtrip_mismatch")
    except (TypeError, ValueError) as exc:
        raise AdapterValidationError("canonical_serialization_failed") from exc


class _AdapterBase:
    dataset_id = ""
    adapter_version = ""

    def __init__(self, root: Path | str):
        self.root = Path(root).resolve()
        self._index: Optional[Dict[str, TableIndexEntry]] = None

    @property
    def adapter_id(self) -> str:
        return f"{self.dataset_id}:{self.adapter_version}"

    def validate_canonical_table(self, artifact: Mapping[str, Any]) -> None:
        _validate_artifact(
            artifact,
            expected_dataset=self.dataset_id,
            expected_adapter_id=self.adapter_id,
        )


class HiTabAdapterV1(_AdapterBase):
    dataset_id = "HiTab"
    adapter_version = "certa_active_v1_hitab_adapter_v1"

    def discover(self) -> Dict[str, Any]:
        index = self.index_tables()
        return {
            "dataset": self.dataset_id,
            "root": str(self.root),
            "table_file_count": len(index),
            "format": "HiTab JSON",
            "table_id_rule": "filename_stem",
        }

    def index_tables(self) -> Dict[str, TableIndexEntry]:
        if self._index is not None:
            return dict(self._index)
        if not self.root.is_dir():
            raise AdapterDiscoveryError(f"hitab_root_missing:{self.root}")
        index: Dict[str, TableIndexEntry] = {}
        for path in sorted(self.root.glob("*.json"), key=lambda item: item.name):
            table_id = path.stem
            if table_id in index:
                raise DuplicateTableIdentityError(
                    f"duplicate_table_identity:{table_id}"
                )
            source_sha256 = _sha256_file(path)
            index[table_id] = TableIndexEntry(
                dataset=self.dataset_id,
                table_id=table_id,
                source_path=path,
                source_sha256=source_sha256,
                source_identity={
                    "table_id_rule": "filename_stem",
                    "native_relative_path": path.name,
                    "native_sha256": source_sha256,
                    "runtime_source_alias": "",
                    "runtime_source_alias_status": (
                        "NOT_AVAILABLE_WITHOUT_RUNTIME_RECORD"
                    ),
                },
            )
        if not index:
            raise AdapterDiscoveryError(f"hitab_table_files_missing:{self.root}")
        self._index = index
        return dict(index)

    def resolve_table(
        self,
        table_id: Any,
        runtime_record: Optional[Mapping[str, Any]] = None,
    ) -> ResolvedNativeTable:
        identity = str(table_id)
        entry = self.index_tables().get(identity)
        if entry is None:
            raise TableResolutionError(f"unresolved_table_id:{identity}")
        try:
            payload = json.loads(entry.source_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise TableResolutionError(
                f"table_parse_failed:{identity}"
            ) from exc
        if not isinstance(payload, dict):
            raise AdapterValidationError(f"native_table_not_object:{identity}")
        source_identity = dict(entry.source_identity)
        if runtime_record is not None:
            alias = str(runtime_record.get("table_source") or "")
            source_identity["runtime_source_alias"] = alias
            source_identity["runtime_source_alias_status"] = (
                "PRESENT_PROVENANCE_ONLY"
                if alias
                else "NOT_PRESENT_IN_AUTHORIZED_RUNTIME_FIELDS"
            )
        return ResolvedNativeTable(
            self.dataset_id,
            identity,
            entry.source_path,
            entry.source_sha256,
            source_identity,
            payload,
        )

    def canonicalize_table(
        self,
        native_table: ResolvedNativeTable,
    ) -> Dict[str, Any]:
        if native_table.dataset != self.dataset_id:
            raise AdapterValidationError("native_dataset_mismatch")
        payload = dict(native_table.native_payload)
        texts = _as_text_rows(payload.get("texts"))
        texts = _pad_hitab_graph_grid(payload, texts)
        graph_payload = dict(payload)
        graph_payload["texts"] = texts
        graph_payload["id"] = native_table.table_id
        graph_payload["table_id"] = native_table.table_id
        graph_payload["top_root"] = _normalize_tree_coordinates(
            payload.get("top_root")
        )
        graph_payload["left_root"] = _normalize_tree_coordinates(
            payload.get("left_root")
        )
        artifact = _build_artifact(
            dataset=self.dataset_id,
            table_id=native_table.table_id,
            adapter_id=self.adapter_id,
            source_identity=native_table.source_identity,
            native_payload=payload,
            graph_payload=graph_payload,
        )
        self.validate_canonical_table(artifact)
        return artifact


def _aitqa_graph_payload(
    raw_table: Mapping[str, Any],
    table_id: str,
) -> Dict[str, Any]:
    column_header = _as_text_rows(raw_table.get("column_header"))
    row_header = _as_text_rows(raw_table.get("row_header"))
    data = _as_text_rows(raw_table.get("data"))
    if not data:
        raise AdapterValidationError("aitqa_data_empty")
    data_columns = max((len(row) for row in data), default=0)
    if data_columns <= 0 or any(len(row) != data_columns for row in data):
        raise AdapterValidationError("aitqa_data_ragged")
    explicit_rows = bool(row_header)
    if explicit_rows and len(row_header) < len(data):
        raise AdapterValidationError("aitqa_row_header_count_mismatch")
    left_columns = (
        max((len(row) for row in row_header), default=0)
        if explicit_rows
        else 1
    )
    if left_columns <= 0:
        raise AdapterValidationError("aitqa_left_header_empty")
    if explicit_rows:
        if len(column_header) == data_columns + left_columns:
            left_header_paths = column_header[:left_columns]
            data_header_paths = column_header[left_columns:]
        elif len(column_header) == data_columns:
            left_header_paths = []
            data_header_paths = column_header
        else:
            raise AdapterValidationError("aitqa_column_header_width_mismatch")
        data_rows = data
        row_paths = row_header[:len(data)]
    else:
        if len(column_header) != data_columns:
            raise AdapterValidationError("aitqa_column_header_width_mismatch")
        if data_columns < 2:
            raise AdapterValidationError("aitqa_no_data_region")
        left_header_paths = column_header[:1]
        data_header_paths = column_header[1:]
        row_paths = [[row[0]] for row in data]
        data_rows = [row[1:] for row in data]
    top_rows = max(
        (len(path) for path in (*left_header_paths, *data_header_paths)),
        default=1,
    )
    top_rows = max(top_rows, 1)
    texts: list[list[str]] = []
    for level in range(top_rows):
        left_values = [
            path[level] if level < len(path) else ""
            for path in left_header_paths
        ]
        left_values.extend([""] * (left_columns - len(left_values)))
        data_values = [
            path[level] if level < len(path) else ""
            for path in data_header_paths
        ]
        texts.append(left_values + data_values)
    for row_path, values in zip(row_paths, data_rows):
        left_values = list(row_path[:left_columns])
        left_values.extend([""] * (left_columns - len(left_values)))
        texts.append(left_values + list(values))
    rows, columns = _rectangular_rows(texts)
    top_root, left_root = _axis_roots(
        rows,
        columns,
        top_rows,
        left_columns,
    )
    return {
        "id": table_id,
        "table_id": table_id,
        "title": str(raw_table.get("title") or raw_table.get("caption") or ""),
        "source_format": "aitqa_raw_structural_v1",
        "texts": texts,
        "top_root": top_root,
        "left_root": left_root,
        "merged_regions": [],
        "top_header_rows_num": top_rows,
        "left_header_columns_num": left_columns,
        "aitqa_meta": {
            "original_column_header_paths": len(column_header),
            "original_row_header_paths": len(row_header),
            "original_data_rows": len(data),
            "has_explicit_row_header": explicit_rows,
            "orphan_row_header_count": max(
                0,
                len(row_header) - len(data),
            ),
        },
    }


class AITQAAdapterV1(_AdapterBase):
    dataset_id = "aitqa"
    adapter_version = "certa_active_v1_aitqa_adapter_v1"

    def __init__(self, root: Path | str):
        super().__init__(root)
        self._payloads: Dict[str, Dict[str, Any]] = {}
        self._line_numbers: Dict[str, list[int]] = {}
        self._clean_tables: Optional[Dict[str, list[list[str]]]] = None

    @property
    def raw_path(self) -> Path:
        return self.root / "test_samples.jsonl"

    @property
    def clean_path(self) -> Path:
        return self.root / "aitqa_clean_questions.json"

    def _load_clean_tables(self) -> Tuple[Dict[str, list[list[str]]], int]:
        if self._clean_tables is not None:
            rows = json.loads(self.clean_path.read_text(encoding="utf-8"))
            return dict(self._clean_tables), len(rows)
        if not self.clean_path.is_file():
            raise AdapterDiscoveryError(
                f"aitqa_clean_questions_missing:{self.clean_path}"
            )
        try:
            rows = json.loads(self.clean_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise AdapterDiscoveryError("aitqa_clean_questions_invalid") from exc
        if not isinstance(rows, list):
            raise AdapterDiscoveryError("aitqa_clean_questions_not_list")
        tables: Dict[str, list[list[str]]] = {}
        for row in rows:
            if not isinstance(row, Mapping):
                raise AdapterDiscoveryError("aitqa_clean_question_not_object")
            table_id = str(row.get("table_id") or "")
            flat = _as_text_rows(row.get("table"))
            if not table_id or not flat:
                raise AdapterDiscoveryError("aitqa_clean_table_identity_empty")
            previous = tables.get(table_id)
            if previous is not None and previous != flat:
                raise DuplicateTableIdentityError(
                    f"clean_table_payload_conflict:{table_id}"
                )
            tables[table_id] = flat
        self._clean_tables = tables
        return dict(tables), len(rows)

    def discover(self) -> Dict[str, Any]:
        index = self.index_tables()
        clean_tables, clean_questions = self._load_clean_tables()
        missing = sorted(set(clean_tables) - set(index))
        if missing:
            raise AdapterDiscoveryError(
                f"aitqa_clean_table_missing_from_raw:{','.join(missing)}"
            )
        for table_id, clean in clean_tables.items():
            graph_payload = _aitqa_graph_payload(
                self._payloads[table_id],
                table_id,
            )
            if graph_payload["texts"] != clean:
                raise AdapterValidationError(
                    f"aitqa_clean_flattening_mismatch:{table_id}"
                )
        return {
            "dataset": self.dataset_id,
            "root": str(self.root),
            "raw_question_count": sum(
                len(lines) for lines in self._line_numbers.values()
            ),
            "raw_table_count": len(index),
            "clean_question_count": clean_questions,
            "clean_table_count": len(clean_tables),
            "raw_path": str(self.raw_path),
            "clean_path": str(self.clean_path),
            "table_id_rule": "raw_row.table_id_equals_raw_table.id",
        }

    def index_tables(self) -> Dict[str, TableIndexEntry]:
        if self._index is not None:
            return dict(self._index)
        if not self.raw_path.is_file():
            raise AdapterDiscoveryError(
                f"aitqa_raw_questions_missing:{self.raw_path}"
            )
        source_sha256 = _sha256_file(self.raw_path)
        index: Dict[str, TableIndexEntry] = {}
        payloads: Dict[str, Dict[str, Any]] = {}
        lines: Dict[str, list[int]] = {}
        with self.raw_path.open("r", encoding="utf-8") as handle:
            for line_number, line in enumerate(handle, start=1):
                if not line.strip():
                    continue
                try:
                    row = json.loads(line)
                except json.JSONDecodeError as exc:
                    raise AdapterDiscoveryError(
                        f"aitqa_raw_json_invalid:{line_number}"
                    ) from exc
                if not isinstance(row, Mapping):
                    raise AdapterDiscoveryError(
                        f"aitqa_raw_row_not_object:{line_number}"
                    )
                raw_table = row.get("table")
                if not isinstance(raw_table, Mapping):
                    raise AdapterValidationError(
                        f"aitqa_raw_table_not_object:{line_number}"
                    )
                row_id = str(row.get("table_id") or "")
                native_id = str(raw_table.get("id") or "")
                if not row_id or row_id != native_id:
                    raise AdapterValidationError(
                        f"aitqa_table_identity_mismatch:{line_number}"
                    )
                payload = dict(raw_table)
                payload_sha256 = canonical_json_hash(payload)
                if row_id in payloads:
                    if canonical_json_hash(payloads[row_id]) != payload_sha256:
                        raise DuplicateTableIdentityError(
                            f"duplicate_table_payload_conflict:{row_id}"
                        )
                    lines[row_id].append(line_number)
                    continue
                payloads[row_id] = payload
                lines[row_id] = [line_number]
                index[row_id] = TableIndexEntry(
                    dataset=self.dataset_id,
                    table_id=row_id,
                    source_path=self.raw_path,
                    source_sha256=source_sha256,
                    source_identity={
                        "table_id_rule": (
                            "raw_row.table_id_equals_raw_table.id"
                        ),
                        "native_relative_path": self.raw_path.name,
                        "native_sha256": source_sha256,
                        "raw_line_numbers": lines[row_id],
                    },
                    payload_sha256=payload_sha256,
                )
        if not index:
            raise AdapterDiscoveryError("aitqa_raw_tables_empty")
        for table_id, entry in tuple(index.items()):
            identity = dict(entry.source_identity)
            identity["raw_line_numbers"] = list(lines[table_id])
            index[table_id] = replace(entry, source_identity=identity)
        self._index = index
        self._payloads = payloads
        self._line_numbers = lines
        return dict(index)

    def resolve_table(
        self,
        table_id: Any,
        runtime_record: Optional[Mapping[str, Any]] = None,
    ) -> ResolvedNativeTable:
        del runtime_record
        identity = str(table_id)
        entry = self.index_tables().get(identity)
        if entry is None:
            raise TableResolutionError(f"unresolved_table_id:{identity}")
        return ResolvedNativeTable(
            self.dataset_id,
            identity,
            entry.source_path,
            entry.source_sha256,
            dict(entry.source_identity),
            dict(self._payloads[identity]),
        )

    def canonicalize_table(
        self,
        native_table: ResolvedNativeTable,
    ) -> Dict[str, Any]:
        if native_table.dataset != self.dataset_id:
            raise AdapterValidationError("native_dataset_mismatch")
        graph_payload = _aitqa_graph_payload(
            native_table.native_payload,
            native_table.table_id,
        )
        clean_tables, _ = self._load_clean_tables()
        if (
            native_table.table_id in clean_tables
            and graph_payload["texts"] != clean_tables[native_table.table_id]
        ):
            raise AdapterValidationError(
                f"aitqa_clean_flattening_mismatch:{native_table.table_id}"
            )
        graph_payload["aitqa_meta"]["clean_flattening_verified"] = (
            native_table.table_id in clean_tables
        )
        artifact = _build_artifact(
            dataset=self.dataset_id,
            table_id=native_table.table_id,
            adapter_id=self.adapter_id,
            source_identity=native_table.source_identity,
            native_payload=native_table.native_payload,
            graph_payload=graph_payload,
        )
        self.validate_canonical_table(artifact)
        return artifact


def _sst_sheet_id(workbook_id: str, index: int, title: str) -> str:
    normalized_title = unicodedata.normalize("NFC", str(title))
    title_sha256 = hashlib.sha256(
        normalized_title.encode("utf-8")
    ).hexdigest()
    return (
        f"sstqa_zh:{workbook_id}:sheet:{index:03d}:"
        f"{title_sha256[:16]}"
    )


def _zero_based_merged_regions(
    sheet: Mapping[str, Any],
) -> list[Dict[str, int]]:
    return [
        {
            "first_row": int(region["first_row"]) - 1,
            "last_row": int(region["last_row"]) - 1,
            "first_column": int(region["first_column"]) - 1,
            "last_column": int(region["last_column"]) - 1,
        }
        for region in (sheet.get("merged_regions") or [])
    ]


def _sst_graph_payload_for_bands(
    *,
    table_id: str,
    sheet: Mapping[str, Any],
    top_header_rows: int,
    left_header_columns: int,
) -> Dict[str, Any]:
    cells = sheet.get("cells")
    if not isinstance(cells, list):
        raise AdapterValidationError("sstqa_sheet_cells_not_list")
    texts = [
        [
            str(cell.get("surface") or "")
            if isinstance(cell, Mapping)
            else ""
            for cell in row
        ]
        for row in cells
    ]
    rows, columns = _rectangular_rows(texts)
    top_root, left_root = _axis_roots(
        rows,
        columns,
        top_header_rows,
        left_header_columns,
    )
    first_text = next(
        (
            text
            for row in texts
            for text in row
            if str(text).strip()
        ),
        "",
    )
    return {
        "id": table_id,
        "table_id": table_id,
        "title": first_text,
        "source_format": "sstqa_zh_sheet_structural_projection_v1",
        "texts": texts,
        "raw_grid": texts,
        "top_root": top_root,
        "left_root": left_root,
        "merged_regions": _zero_based_merged_regions(sheet),
        "top_header_rows_num": top_header_rows,
        "left_header_columns_num": left_header_columns,
        "sstqa_zh_meta": {
            "sheet_title": str(sheet.get("name") or ""),
            "structural_projection": {
                "algorithm": SST_STRUCTURAL_PROJECTION_VERSION,
                "semantic_header_claim": False,
                "top_header_rows": top_header_rows,
                "left_header_columns": left_header_columns,
            },
        },
    }


@dataclass(frozen=True)
class _GroundingWitness:
    node_id: str
    row: int
    column: int
    row_binding_ids: Tuple[str, ...]
    column_binding_ids: Tuple[str, ...]


def _grounding_witness(graph: HCEG) -> Optional[_GroundingWitness]:
    cells = []
    for node in sorted(
        graph.nodes.values(),
        key=lambda item: (item.row, item.col, item.node_id),
    ):
        if node.node_type != NodeType.CELL:
            continue
        row_bindings = tuple(
            sorted(
                {
                    edge.target
                    for edge in graph._adj.get(node.node_id, ())
                    if edge.edge_type == EdgeType.ROW_PATH
                }
            )
        )
        column_bindings = tuple(
            sorted(
                {
                    edge.target
                    for edge in graph._adj.get(node.node_id, ())
                    if edge.edge_type
                    in {EdgeType.COL_PATH, EdgeType.VALUE_UNDER_HEADER}
                }
            )
        )
        cells.append((node, row_bindings, column_bindings))
    binding_sets = [
        set((*row_bindings, *column_bindings))
        for _, row_bindings, column_bindings in cells
    ]
    for index, (node, row_bindings, column_bindings) in enumerate(cells):
        if not str(node.text or "").strip() or not row_bindings or not column_bindings:
            continue
        wanted = binding_sets[index]
        if sum(wanted.issubset(other) for other in binding_sets) != 1:
            continue
        resolution = resolve_atomic_operand(graph, wanted)
        if (
            resolution.state == ResolutionState.UNIQUE
            and resolution.unique_node_id == node.node_id
        ):
            return _GroundingWitness(
                node.node_id,
                int(node.row),
                int(node.col),
                row_bindings,
                column_bindings,
            )
    return None


def _sst_graph_payload(
    table_id: str,
    sheet: Mapping[str, Any],
) -> Dict[str, Any]:
    cells = sheet.get("cells")
    if not isinstance(cells, list) or not cells:
        raise AdapterValidationError("empty_sheet")
    texts = [
        [
            str(cell.get("surface") or "")
            if isinstance(cell, Mapping)
            else ""
            for cell in row
        ]
        for row in cells
    ]
    rows, columns = _rectangular_rows(texts)
    populated = [
        (row_index, column_index)
        for row_index, row in enumerate(texts)
        for column_index, text in enumerate(row)
        if str(text).strip()
    ]
    if not populated:
        raise AdapterValidationError("empty_sheet")
    minimum_top = min(row for row, _ in populated) + 1
    minimum_left = min(column for _, column in populated) + 1
    if minimum_top >= rows or minimum_left >= columns:
        raise AdapterValidationError("sstqa_sheet_has_no_data_region")
    for total in range(minimum_top + minimum_left, rows + columns):
        for top_header_rows in range(minimum_top, rows):
            left_header_columns = total - top_header_rows
            if not (
                minimum_left <= left_header_columns < columns
            ):
                continue
            payload = _sst_graph_payload_for_bands(
                table_id=table_id,
                sheet=sheet,
                top_header_rows=top_header_rows,
                left_header_columns=left_header_columns,
            )
            graph = build_hceg(
                payload,
                add_spatial=False,
                add_value_nodes=False,
            )
            if _grounding_witness(graph) is not None:
                return payload
    raise AdapterValidationError(
        "sstqa_no_exact_groundable_structural_projection"
    )


class SSTQAZhAdapterV1(_AdapterBase):
    dataset_id = "sstqa_zh"
    adapter_version = "certa_active_v1_sstqa_zh_adapter_v1"

    def __init__(self, root: Path | str):
        super().__init__(root)
        self.workbook_root = (
            self.root / "table"
            if (self.root / "table").is_dir()
            else self.root
        )
        self._active_aliases: Dict[str, str] = {}

    def discover(self) -> Dict[str, Any]:
        index = self.index_tables()
        return {
            "dataset": self.dataset_id,
            "root": str(self.root),
            "workbook_root": str(self.workbook_root),
            "workbook_count": len(self._active_aliases),
            "sheet_count": len(index),
            "format": "XLSX workbook sheets",
            "table_id_rule": (
                "sstqa_zh:{workbook_id}:sheet:{index:03d}:"
                "sha256(NFC(title))[:16]"
            ),
            "question_alias_rule": "workbook_id_to_active_sheet",
        }

    def index_tables(self) -> Dict[str, TableIndexEntry]:
        if self._index is not None:
            return dict(self._index)
        if not self.workbook_root.is_dir():
            raise AdapterDiscoveryError(
                f"sstqa_workbook_root_missing:{self.workbook_root}"
            )
        index: Dict[str, TableIndexEntry] = {}
        aliases: Dict[str, str] = {}
        paths = sorted(
            self.workbook_root.glob("*.xlsx"),
            key=lambda item: item.name,
        )
        if not paths:
            raise AdapterDiscoveryError("sstqa_workbooks_missing")
        for path in paths:
            workbook_id = path.stem
            if workbook_id in aliases:
                raise DuplicateTableIdentityError(
                    f"duplicate_workbook_identity:{workbook_id}"
                )
            workbook_sha256 = _sha256_file(path)
            try:
                workbook = openpyxl.load_workbook(
                    path,
                    data_only=False,
                    read_only=False,
                )
            except (OSError, ValueError) as exc:
                raise AdapterDiscoveryError(
                    f"sstqa_workbook_parse_failed:{workbook_id}"
                ) from exc
            try:
                active_index = workbook.index(workbook.active)
                for sheet_index, sheet in enumerate(workbook.worksheets):
                    title = str(sheet.title)
                    normalized_title = unicodedata.normalize("NFC", title)
                    title_sha256 = hashlib.sha256(
                        normalized_title.encode("utf-8")
                    ).hexdigest()
                    table_id = _sst_sheet_id(
                        workbook_id,
                        sheet_index,
                        title,
                    )
                    if table_id in index:
                        raise DuplicateTableIdentityError(
                            f"duplicate_table_identity:{table_id}"
                        )
                    identity = {
                        "table_id_rule": (
                            "workbook_id+sheet_index+NFC_title_sha256"
                        ),
                        "native_relative_path": path.name,
                        "workbook_id": workbook_id,
                        "workbook_sha256": workbook_sha256,
                        "sheet_index": sheet_index,
                        "sheet_title": title,
                        "sheet_title_nfc": normalized_title,
                        "sheet_title_sha256": title_sha256,
                        "sheet_state": str(sheet.sheet_state),
                        "is_active_sheet": sheet_index == active_index,
                        "native_rows": int(sheet.max_row),
                        "native_columns": int(sheet.max_column),
                        "native_merged_region_count": len(
                            sheet.merged_cells.ranges
                        ),
                    }
                    index[table_id] = TableIndexEntry(
                        dataset=self.dataset_id,
                        table_id=table_id,
                        source_path=path,
                        source_sha256=workbook_sha256,
                        source_identity=identity,
                    )
                    if sheet_index == active_index:
                        aliases[workbook_id] = table_id
            finally:
                workbook.close()
        self._index = index
        self._active_aliases = aliases
        return dict(index)

    def resolve_table(
        self,
        table_id: Any,
        runtime_record: Optional[Mapping[str, Any]] = None,
    ) -> ResolvedNativeTable:
        del runtime_record
        identity = str(table_id)
        index = self.index_tables()
        resolved_id = self._active_aliases.get(identity, identity)
        entry = index.get(resolved_id)
        if entry is None:
            raise TableResolutionError(f"unresolved_table_id:{identity}")
        workbook_id = str(entry.source_identity["workbook_id"])
        workbook = convert_sstqa_workbook(
            entry.source_path,
            table_id=workbook_id,
        )
        sheet_index = int(entry.source_identity["sheet_index"])
        sheets = workbook.get("sheets")
        if not isinstance(sheets, list) or sheet_index >= len(sheets):
            raise AdapterValidationError(
                f"sstqa_sheet_index_missing:{resolved_id}"
            )
        sheet = sheets[sheet_index]
        if str(sheet.get("name") or "") != entry.source_identity["sheet_title"]:
            raise AdapterValidationError(
                f"sstqa_sheet_title_mismatch:{resolved_id}"
            )
        native_payload = {
            "schema_version": str(workbook.get("schema_version") or ""),
            "source_format": str(workbook.get("source_format") or ""),
            "source_file": str(workbook.get("source_file") or ""),
            "source_sha256": str(workbook.get("source_sha256") or ""),
            "workbook_id": workbook_id,
            "workbook_active_sheet_id": self._active_aliases[workbook_id],
            "sheet_id": resolved_id,
            "sheet": dict(sheet),
        }
        return ResolvedNativeTable(
            self.dataset_id,
            resolved_id,
            entry.source_path,
            entry.source_sha256,
            dict(entry.source_identity),
            native_payload,
        )

    def canonicalize_table(
        self,
        native_table: ResolvedNativeTable,
    ) -> Dict[str, Any]:
        if native_table.dataset != self.dataset_id:
            raise AdapterValidationError("native_dataset_mismatch")
        sheet = native_table.native_payload.get("sheet")
        if not isinstance(sheet, Mapping):
            raise AdapterValidationError("sstqa_sheet_payload_missing")
        graph_payload = _sst_graph_payload(
            native_table.table_id,
            sheet,
        )
        artifact = _build_artifact(
            dataset=self.dataset_id,
            table_id=native_table.table_id,
            adapter_id=self.adapter_id,
            source_identity=native_table.source_identity,
            native_payload=native_table.native_payload,
            graph_payload=graph_payload,
        )
        self.validate_canonical_table(artifact)
        return artifact


def _projected_value_matches(node: Any, projected_answer: str) -> bool:
    if str(projected_answer) == str(node.text or ""):
        return True
    if node.numeric_value is not None:
        try:
            numeric = (
                str(projected_answer)
                .replace(",", "")
                .replace("$", "")
                .replace("¥", "")
                .replace("%", "")
                .strip()
            )
            return float(numeric) == float(node.numeric_value)
        except (TypeError, ValueError):
            return False
    text = str(node.text or "")
    if text.strip().lower() in {"true", "false"}:
        return str(projected_answer).strip().lower() == text.strip().lower()
    return str(projected_answer) == text


def roundtrip_adapter_artifact(
    artifact: Mapping[str, Any],
    *,
    question: str,
) -> Dict[str, Any]:
    """Exercise canonical table → graph → Planner → local exact execution."""
    _validate_artifact(artifact)
    graph_payload = artifact["table_payload"]["graph_payload"]
    graph = build_hceg(
        dict(graph_payload),
        str(question or ""),
        add_spatial=False,
        add_value_nodes=True,
    )
    planner_view = build_proposal_blind_planner_view(
        question=str(question or ""),
        graph=graph,
        table_json=graph_payload,
        query_contract=None,
        include_table_values=False,
        legacy_query_semantics_mode="audit_only",
    )
    groups = build_canonical_structural_group_catalog(
        graph=graph,
        table_json=graph_payload,
    )
    witness = _grounding_witness(graph)
    if witness is None:
        return {
            "pass": False,
            "failure_reasons": ["no_exact_grounding_witness"],
            "graph_node_count": len(graph.nodes),
            "planner_schema_node_count": len(planner_view["schema_nodes"]),
            "planner_schema_edge_count": len(planner_view["schema_edges"]),
            "structural_group_count": len(groups["all_groups"]),
            "grounding_state": "UNRESOLVED",
            "closure_outcome": "NOT_RUN",
        }
    node = graph.nodes[witness.node_id]
    if node.numeric_value is not None:
        signature_id = "LOOKUP_VALUE_SCALAR"
        answer_domain = "SCALAR"
    elif str(node.text or "").strip().lower() in {"true", "false"}:
        signature_id = "LOOKUP_VALUE_BOOLEAN"
        answer_domain = "BOOLEAN"
    else:
        signature_id = "LOOKUP_VALUE_ENTITY"
        answer_domain = "ENTITY"
    resolution = resolve_atomic_operand(
        graph,
        (*witness.row_binding_ids, *witness.column_binding_ids),
    )
    plan = {
        "plan_id": "ADAPTER_SMOKE_LOOKUP",
        "signature_id": signature_id,
        "operation_family": "LOOKUP",
        "semantic_result_role": "VALUE",
        "projection_operator": "VALUE_PROJECTION",
        "answer_domain": answer_domain,
        "role_bindings": {
            "TARGET_ENTITY": list(witness.row_binding_ids),
            "TARGET_MEASURE": list(witness.column_binding_ids),
        },
        "unresolved_semantics": [],
    }
    closure = build_plan_closure(
        {
            "planner_version": "certa_adapter_roundtrip_v1",
            "plans": [plan],
        },
        graph,
        allowed_signature_ids=[signature_id],
    )
    assignment = closure.assignments[0] if closure.assignments else None
    derivation = (
        closure.executable_derivations[0]
        if closure.executable_derivations
        else None
    )
    native_coordinate = ""
    native_payload = artifact["table_payload"]["native_payload"]
    sheet = native_payload.get("sheet")
    if isinstance(sheet, Mapping):
        cells = sheet.get("cells")
        if (
            isinstance(cells, list)
            and witness.row < len(cells)
            and isinstance(cells[witness.row], list)
            and witness.column < len(cells[witness.row])
            and isinstance(cells[witness.row][witness.column], Mapping)
        ):
            native_coordinate = str(
                cells[witness.row][witness.column].get("coordinate") or ""
            )
    projection_match = bool(
        derivation is not None
        and witness.node_id in derivation.operand_node_ids
        and _projected_value_matches(node, derivation.projected_answer)
    )
    failure_reasons = []
    if not graph.nodes:
        failure_reasons.append("graph_empty")
    if not planner_view["schema_nodes"]:
        failure_reasons.append("planner_schema_empty")
    if not groups["all_groups"]:
        failure_reasons.append("structural_groups_empty")
    if resolution.state != ResolutionState.UNIQUE:
        failure_reasons.append("grounding_not_unique")
    if assignment is None or assignment.outcome != ClosureOutcome.UNIQUE_EXECUTABLE:
        failure_reasons.append("closure_not_unique_executable")
    if derivation is None:
        failure_reasons.append("derivation_missing")
    elif not derivation.provenance_complete:
        failure_reasons.append("derivation_provenance_incomplete")
    if not projection_match:
        failure_reasons.append("projection_not_equal_to_grounded_cell")
    return {
        "pass": not failure_reasons,
        "failure_reasons": failure_reasons,
        "graph_node_count": len(graph.nodes),
        "graph_edge_count": len(graph.edges),
        "planner_schema_node_count": len(planner_view["schema_nodes"]),
        "planner_schema_edge_count": len(planner_view["schema_edges"]),
        "structural_group_count": len(groups["all_groups"]),
        "structural_group_catalog_sha256": groups["catalog_sha256"],
        "grounding_state": resolution.state.value,
        "grounded_node_id": witness.node_id,
        "graph_row": witness.row,
        "graph_column": witness.column,
        "native_coordinate": native_coordinate,
        "row_binding_ids": list(witness.row_binding_ids),
        "column_binding_ids": list(witness.column_binding_ids),
        "closure_outcome": (
            assignment.outcome.value if assignment is not None else "NOT_RUN"
        ),
        "executable_derivation_count": len(
            closure.executable_derivations
        ),
        "projected_answer": (
            derivation.projected_answer if derivation is not None else ""
        ),
        "projection_matches_grounded_cell": projection_match,
    }
