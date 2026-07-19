"""Deterministic, layout-preserving SSTQA_zh workbook support."""

from __future__ import annotations

import datetime as dt
import hashlib
import math
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional, Tuple

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries


SCHEMA_VERSION = "certa_sstqa_zh_table_v1"
HEADER_INFERENCE_VERSION = "none_raw_grid_v1"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _json_value(value: Any) -> Any:
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return str(value)
    return value


def _decimal_places(number_format: str) -> int:
    positive = number_format.split(";", 1)[0]
    match = re.search(r"\.([0#]+)", positive)
    return len(match.group(1)) if match else 0


def _numeric_surface(value: float, number_format: str) -> str:
    if number_format == "General":
        return str(int(value)) if float(value).is_integer() else format(value, ".15g")
    positive = number_format.split(";", 1)[0]
    decimals = _decimal_places(positive)
    percent = "%" in positive
    rendered_value = value * 100 if percent else value
    grouping = "#,##" in positive
    rendered = f"{rendered_value:,.{decimals}f}" if grouping else f"{rendered_value:.{decimals}f}"
    prefix = ""
    if "¥" in positive or "￥" in positive:
        prefix = "¥"
    elif "$" in positive:
        prefix = "$"
    return f"{prefix}{rendered}{'%' if percent else ''}"


def _cell_surface(value: Any, number_format: str, *, formula: str = "") -> str:
    if formula and value is None:
        return formula
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return _numeric_surface(float(value), number_format)
    return str(value)


def _merged_anchors(sheet: Any) -> Tuple[Dict[str, str], List[Dict[str, Any]]]:
    anchors: Dict[str, str] = {}
    regions = []
    for merged_range in sorted(sheet.merged_cells.ranges, key=lambda item: str(item)):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        anchor = f"{get_column_letter(min_col)}{min_row}"
        regions.append(
            {
                "range": str(merged_range),
                "first_row": min_row,
                "last_row": max_row,
                "first_column": min_col,
                "last_column": max_col,
                "anchor": anchor,
            }
        )
        for row in range(min_row, max_row + 1):
            for column in range(min_col, max_col + 1):
                anchors[f"{get_column_letter(column)}{row}"] = anchor
    return anchors, regions


def _sheet_payload(formula_sheet: Any, value_sheet: Any) -> Dict[str, Any]:
    anchors, merged_regions = _merged_anchors(formula_sheet)
    cells = []
    nonempty_count = 0
    for row_index in range(1, formula_sheet.max_row + 1):
        row_payload = []
        for column_index in range(1, formula_sheet.max_column + 1):
            formula_cell = formula_sheet.cell(row=row_index, column=column_index)
            value_cell = value_sheet.cell(row=row_index, column=column_index)
            coordinate = f"{get_column_letter(column_index)}{row_index}"
            is_merged_child = isinstance(formula_cell, MergedCell)
            raw_formula = ""
            raw_value = None
            cached_value = None
            data_type = "n"
            number_format = "General"
            style_id = 0
            if not is_merged_child:
                data_type = str(formula_cell.data_type)
                number_format = str(formula_cell.number_format or "General")
                style_id = int(formula_cell.style_id)
                if data_type == "f":
                    raw_formula = str(formula_cell.value or "")
                    cached_value = value_cell.value
                else:
                    raw_value = formula_cell.value
                    cached_value = formula_cell.value
            surface_value = cached_value if raw_formula and cached_value is not None else raw_value
            surface = _cell_surface(surface_value, number_format, formula=raw_formula)
            if raw_value is not None or raw_formula:
                nonempty_count += 1
            row_payload.append(
                {
                    "row": row_index,
                    "column": column_index,
                    "coordinate": coordinate,
                    "raw_value": _json_value(raw_value),
                    "cached_value": _json_value(cached_value),
                    "formula": raw_formula,
                    "surface": surface,
                    "data_type": data_type,
                    "number_format": number_format,
                    "style_class": f"style_{style_id}",
                    "merged_anchor": anchors.get(coordinate, ""),
                    "is_merged_anchor": anchors.get(coordinate) == coordinate,
                }
            )
        cells.append(row_payload)
    row_heights = {
        str(index): float(dimension.height)
        for index, dimension in sorted(formula_sheet.row_dimensions.items())
        if dimension.height is not None
    }
    column_widths = {
        str(column): float(dimension.width)
        for column, dimension in sorted(formula_sheet.column_dimensions.items())
        if dimension.width is not None
    }
    return {
        "name": formula_sheet.title,
        "state": formula_sheet.sheet_state,
        "used_range": formula_sheet.calculate_dimension(),
        "row_count": formula_sheet.max_row,
        "column_count": formula_sheet.max_column,
        "nonempty_cell_count": nonempty_count,
        "cells": cells,
        "merged_regions": merged_regions,
        "row_heights": row_heights,
        "column_widths": column_widths,
    }


def convert_sstqa_workbook(path: Path, *, table_id: Optional[str] = None) -> Dict[str, Any]:
    path = Path(path)
    formulas = openpyxl.load_workbook(path, data_only=False, read_only=False)
    values = openpyxl.load_workbook(path, data_only=True, read_only=False)
    value_sheets = {sheet.title: sheet for sheet in values.worksheets}
    sheets = [_sheet_payload(sheet, value_sheets[sheet.title]) for sheet in formulas.worksheets]
    if not sheets:
        raise ValueError(f"SSTQA_zh workbook has no worksheets: {path}")
    primary = max(sheets, key=lambda sheet: (sheet["nonempty_cell_count"], -sheets.index(sheet)))
    return {
        "schema_version": SCHEMA_VERSION,
        "table_id": str(table_id if table_id is not None else path.stem),
        "source_format": "sstqa_zh_xlsx",
        "source_file": path.name,
        "source_sha256": _sha256(path),
        "primary_sheet": primary["name"],
        "sheet_count": len(sheets),
        "header_inference": HEADER_INFERENCE_VERSION,
        "sheets": sheets,
    }


def _root_node() -> Dict[str, Any]:
    return {"row": -1, "column": -1, "row_index": -1, "column_index": -1, "children": []}


def canonical_table_to_hitab_like(table: Mapping[str, Any]) -> Dict[str, Any]:
    sheets = list(table.get("sheets") or [])
    primary_name = str(table.get("primary_sheet", ""))
    primary = next((sheet for sheet in sheets if sheet.get("name") == primary_name), sheets[0] if sheets else {})
    cells = primary.get("cells") or []
    texts = [[str(cell.get("surface", "")) for cell in row] for row in cells]
    merged_regions = [
        {
            "first_row": int(region["first_row"]) - 1,
            "last_row": int(region["last_row"]) - 1,
            "first_column": int(region["first_column"]) - 1,
            "last_column": int(region["last_column"]) - 1,
        }
        for region in (primary.get("merged_regions") or [])
    ]
    table_id = str(table.get("table_id", ""))
    return {
        "id": table_id,
        "table_id": table_id,
        "title": texts[0][0] if texts and texts[0] else "",
        "source_format": "sstqa_zh",
        "texts": texts,
        "raw_grid": texts,
        "top_root": _root_node(),
        "left_root": _root_node(),
        "merged_regions": merged_regions,
        "top_header_rows_num": 0,
        "left_header_columns_num": 0,
        "sstqa_zh_meta": {
            "schema_version": table.get("schema_version", ""),
            "source_sha256": table.get("source_sha256", ""),
            "primary_sheet": primary_name,
            "sheet_count": table.get("sheet_count", len(sheets)),
            "header_inference": table.get("header_inference", HEADER_INFERENCE_VERSION),
            "sheets": sheets,
        },
    }


def serialize_canonical_grid(table: Mapping[str, Any], *, max_characters: int) -> Dict[str, Any]:
    if max_characters < 0:
        raise ValueError("max_characters must be nonnegative")
    candidates = []
    for sheet in table.get("sheets") or []:
        sheet_name = str(sheet.get("name", ""))
        for row in sheet.get("cells") or []:
            for cell in row:
                surface = str(cell.get("surface", ""))
                if not surface:
                    continue
                merged_anchor = str(cell.get("merged_anchor", ""))
                if merged_anchor and not cell.get("is_merged_anchor", False):
                    continue
                identity = f"{sheet_name}!{cell.get('coordinate', '')}"
                candidates.append((identity, f"[{identity}] {surface}"))

    lines = []
    included = []
    for identity, line in candidates:
        proposed = "\n".join(lines + [line])
        if len(proposed) > max_characters:
            break
        lines.append(line)
        included.append(identity)
    text = "\n".join(lines)
    return {
        "text": text,
        "included_cells": included,
        "character_count": len(text),
        "candidate_cell_count": len(candidates),
        "truncated": len(included) < len(candidates),
    }


_SURROUNDING_QUOTES = (("\"", "\""), ("'", "'"), ("“", "”"), ("‘", "’"))
_LIST_SEPARATOR = re.compile(r"[,;、]+")
_NUMBER_WITH_UNIT = re.compile(
    r"^(?P<prefix>[¥￥$])?(?P<number>[-+]?(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?)(?P<unit>[^\d]*)$"
)


def normalize_sstqa_text(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    changed = True
    while changed and len(text) >= 2:
        changed = False
        for left, right in _SURROUNDING_QUOTES:
            if text.startswith(left) and text.endswith(right):
                text = text[len(left) : len(text) - len(right)].strip()
                changed = True
                break
    return "".join(text.split())


def _number_and_unit(value: Any) -> Optional[Tuple[str, str]]:
    text = normalize_sstqa_text(value)
    match = _NUMBER_WITH_UNIT.fullmatch(text)
    if not match:
        return None
    number_text = match.group("number").replace(",", "")
    try:
        number = float(number_text)
    except ValueError:
        return None
    canonical_number = str(int(number)) if number.is_integer() else format(number, ".15g")
    unit = f"{match.group('prefix') or ''}{match.group('unit') or ''}"
    return canonical_number, unit


def _list_view(value: Any) -> Tuple[str, ...]:
    text = normalize_sstqa_text(value)
    if not _LIST_SEPARATOR.search(text):
        return ()
    return tuple(part for part in (_LIST_SEPARATOR.split(text)) if part)


def evaluate_sstqa_answer(prediction: Any, gold: Any) -> Dict[str, bool]:
    prediction_text = normalize_sstqa_text(prediction)
    gold_text = normalize_sstqa_text(gold)
    text_exact = bool(prediction_text) and prediction_text == gold_text

    prediction_number = _number_and_unit(prediction)
    gold_number = _number_and_unit(gold)
    numeric_match = bool(
        prediction_number
        and gold_number
        and not prediction_number[1]
        and not gold_number[1]
        and prediction_number[0] == gold_number[0]
    )
    numeric_unit_match = bool(
        prediction_number
        and gold_number
        and prediction_number[1]
        and gold_number[1]
        and prediction_number == gold_number
    )

    prediction_list = _list_view(prediction)
    gold_list = _list_view(gold)
    ordered_list_match = bool(prediction_list and gold_list and prediction_list == gold_list)
    unordered_set_match = bool(prediction_list and gold_list and sorted(set(prediction_list)) == sorted(set(gold_list)))
    official_match = bool(text_exact or numeric_match or numeric_unit_match or ordered_list_match)
    return {
        "text_exact": text_exact,
        "numeric_match": numeric_match,
        "numeric_unit_match": numeric_unit_match,
        "ordered_list_match": ordered_list_match,
        "unordered_set_match": unordered_set_match,
        "official_match": official_match,
    }
